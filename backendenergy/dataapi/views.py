from .models import EnergyCost, EnergyTradingCompany
from .serializers import EnergyCostExcelSheet, EnergyCostSerializer, EnergyCompanySerializer
from rest_framework import viewsets, permissions, response, generics
from rest_framework.parsers import MultiPartParser
from openpyxl import load_workbook
from django.utils.datastructures import MultiValueDictKeyError 
from django.views.decorators.csrf import csrf_exempt
from datetime import date, datetime
import asyncio
from concurrent.futures import ThreadPoolExecutor

# Pools for non-blocking IO
executor = ThreadPoolExecutor(
        max_workers=3,
)

# New event loop
event_loop = asyncio.new_event_loop()

# Function to process spreedsheet and add to db
def upload_excel_sheet_to_db(file):
    wb = load_workbook(file)
    print("Hey, I'm running!")
    
    energy_costs_to_add = []
    energy_costs_to_update = []
    
    for sheet in wb.worksheets:

        trading_company, created = EnergyTradingCompany.objects.get_or_create(sheet_name = sheet.title)
        energy_costs = EnergyCost.objects.filter(trading_company=trading_company)

        for row in sheet.iter_rows(min_row=2, max_row=13):
            month = sheet['A'].index(row[0])
            years = [cell.value for cell in sheet[1][1:]]
            values = [cell.value for cell in row[1:]]
                
            for year, value in zip(years, values):
                energy_cost = EnergyCost(
                    month_year = date(month=month, year=int(year), day=1),
                    cost = value,
                    trading_company = trading_company 
                )

                try:
                    to_add = energy_costs.get(month_year = energy_cost.month_year)
                    energy_costs_to_update.append(to_add)
                except EnergyCost.DoesNotExist:
                    energy_costs_to_add.append(energy_cost)
                    
    if len(energy_costs_to_add)>0:
        EnergyCost.objects.bulk_create(energy_costs_to_add)
    
    if len(energy_costs_to_update)>0:
        EnergyCost.objects.bulk_update(energy_costs_to_update, ['cost'] )

    print("Done!")


# Create your views here.
class EnergyTradingCompanyViewSet(viewsets.ModelViewSet):
    queryset = EnergyTradingCompany.objects.all()
    serializer_class = EnergyCompanySerializer
    permission_classes = [permissions.AllowAny]

class EnergyCostViewSet(viewsets.ModelViewSet):
    queryset = EnergyCost.objects.all()
    lookup_field = 'company_id'
    lookup_value_regex = '[\d]+'
    serializer_class = EnergyCostSerializer
    permission_classes = [permissions.AllowAny]

    def get_queryset(self):
        sdate = self.request.query_params.get("sdate", None)
        edate = self.request.query_params.get("edate", None)
        queryset = self.queryset.order_by('month_year')

        if sdate is not None:
            queryset = queryset.filter(month_year__gte = datetime.fromtimestamp(sdate/1000))
        
        if edate is not None:
            queryset = queryset.filter(month_year__lte = datetime.fromtimestamp(edate/1000))
        

        return queryset

    def retrieve(self, request, company_id=None):
        queryset = self.get_queryset().all()
        queryset = queryset.filter(trading_company=EnergyTradingCompany.objects.get(id=company_id))
        
        count = self.request.query_params.get("count", None)

        if count is not None:
            queryset = queryset.reverse()[:int(count)] 

        serializer = EnergyCostSerializer(queryset, many=True)
        return response.Response([[cost["unix_timestamp"]*1000, cost["cost"]] for cost in serializer.data])


    def list(self, request):
        trading_companies = EnergyTradingCompany.objects.all()
        count = self.request.query_params.get("count", None)

        if count is not None:
            count = int(count)


        ret = []

        for trading_company in trading_companies:

            queryset = self.get_queryset().filter(trading_company=trading_company).order_by('month_year')
            queryset = queryset.reverse()[:count].reverse() if count is not None else queryset

            serializer = EnergyCostSerializer(queryset, many=True)

            ret += [{
                'trading_company': trading_company.id,
                'energy_cost': [
                    (
                        serialized_cost['unix_timestamp']*1000, 
                        serialized_cost['cost']
                    ) for serialized_cost in serializer.data
                ],
            }]     

        return response.Response(ret)
    



class UploadXlsx(generics.CreateAPIView):
    serializer_class = EnergyCostExcelSheet
    parser_classes = [MultiPartParser]

    def post(self, request, filename, format=None):
        try:
            file = request.FILES['file']

            if file.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                print("Go to executor!")
                blocking_task = event_loop.run_in_executor(
                    executor,
                    upload_excel_sheet_to_db,
                    file
                )

            return response.Response({
                'content_type': request.FILES['file'].content_type,
                'data': str(request.data)
            })

        except MultiValueDictKeyError as e:
            print(request.FILES)
            return response.Response({
                'detail': "MultiValueDictKeyError"
            })
